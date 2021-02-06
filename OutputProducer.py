from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font
from openpyxl.chart import BarChart,BarChart3D,Reference,Series,layout
from openpyxl.chart.marker import DataPoint
from StringComparator import StringComparator
import  time
from QuizPart import QuizPart
from Question import Question
from Student import Student
#singleton pattern implemented
class OutputProducer:
    __instance = None

    def __init__(self):
        raise RuntimeError('Call instance() method instead!')

    @classmethod
    def instance(cls):
        if cls.__instance is None:
            cls.__instance = cls.__new__(cls)
        return cls.__instance

    def addIntoExecutionLog(self, logInfo):
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")

        f = open("logFile.txt", "a")
        f.write(dt_string + " " + logInfo + '\n')
        f.close()

        print(dt_string + " " + logInfo + '\n')



    def printAttendenceReport(self,studentList):

        global qlist, attPer
        attdatalist = []

        for stu in studentList:
            name = stu.getName().capitalize()
            surname = stu.getSurname().capitalize()
            id = stu.getStudentId()

            if stu.getAttendence == 0 and stu.getNumberOfClasses == 0:
                attRate = 0
            else:
                try:
                    attRate = str(stu.getAttendence().getNumStuAttend()) + "/" + str(
                        stu.getAttendence().getNumClasses())
                    attPer = 100 * stu.getAttendence().getNumStuAttend() / stu.getAttendence().getNumClasses()
                    attPer = float("{:.2f}".format(attPer))
                except:
                    attRate = ""
                    attPer = ""
                    pass

            numPoll = len(stu.getQuizes())

            data = {"Student Id": id, "Name": name, "Surname": surname,
                    "Attendence": stu.getAttendence().getNumStuAttend(),
                    "Num Classes": stu.getAttendence().getNumClasses(), "Attendance Rate": attRate,
                    "Attendence Percentage": attPer, "Number Of Polls": numPoll}
            attdatalist.append(data)

        df = pd.DataFrame(attdatalist)
        df.to_excel("./attendence_results/attendence_report.xlsx")
        self.addIntoExecutionLog("Attendence report is generated.")

    def printStudentOverallResults(self,studentList,poll_list):

        sum=0
        for poll in poll_list:
            sum+=len(poll_list[poll].quizParts)


        overallResults=[]
        for i, stu in enumerate(studentList):
            name=stu.getName().capitalize()
            surname=stu.getSurname().capitalize()
            id=stu.getStudentId()
            totalNumCorrect=stu.numTotalCorrect
            totalNumWrong=stu.numTotalWrong

            data={"Student Id":id, "Name":name,"Surname":surname,"Correct Answer Rate":str(totalNumCorrect)+"/"+str(sum),"Wrong Answer Rate":str(totalNumWrong)+"/"+str(sum)}
            overallResults.append(data)
        
        overallResults=sorted(overallResults,key=lambda i: i["Correct Answer Rate"])
        pd.DataFrame(overallResults).to_excel("./poll_results/Total Success Rates.xlsx")

    def printPollResults(self,studentList,poll_List):
        pollResults = {}

        for i, key in enumerate(poll_List):

            quizDataList = []

            for j,stu in enumerate(studentList):
                name = stu.getName().capitalize()
                surname = stu.getSurname().capitalize()
                id = stu.getStudentId()
                data = {"Student Id": id, "Name": name, "Surname": surname}

                try:

                    exist=False
                    for quiz in stu.getQuizes():
                        if quiz.getQuizName()==key:
                            exist=True
                            qlist = quiz.getQuizParts()
                            for q in qlist:
                                data["q" + str(q.getQuestion().getQuestionNumber())] = q.getIsCorrect()

                            data["Number Of Questions"] = len(qlist)
                            data["Success Rate"] = str(quiz.getNumCorrect()) + "/" + str(len(qlist))
                            data["Success Percentage"] = 100 * quiz.getNumCorrect() / len(qlist)
                            break

                    if exist==False:

                        for x in range(1,poll_List[key].numberOfQuestions + 1):
                            data["q" + str(x)] = " "

                    quizDataList.append(data)


                except:
                    pass
            quiz_df = pd.DataFrame(quizDataList)
            pollResults[key] = quiz_df
            i += 1

        for key in pollResults:
            pollResults[key].to_excel("./poll_results/" + key + "_result.xlsx")

    def printPollStat(self,quizStats):


        for i,keyQuizName in enumerate(quizStats):


            writer = pd.ExcelWriter("./poll_results/" + keyQuizName + "_stats.xlsx", engine='xlsxwriter')

            quizStat=quizStats[keyQuizName]



            for j, quizPart in enumerate(quizStat.questionStatDict):
                questionStat=quizStat.questionStatDict[quizPart]
                answerNumDict=questionStat.getAnswerStat().answerNumDict

                dfAnswers = pd.DataFrame(answerNumDict, columns=answerNumDict.keys(), index=[0])
                dfAnswers = dfAnswers.transpose().rename(columns={0: "count", 1: "%"})

                nums = [n for n in answerNumDict.values()]

                mysum = sum(nums)

                for key in answerNumDict:
                    dfAnswers.loc[key, '%'] = float("{:.2f}".format(100 * (answerNumDict[key] / mysum)))

                dfAnswers.index.name=quizPart.getQuestion().getQuestionText()
                dfAnswers.to_excel(writer, sheet_name="q" + str(j+1))


            writer.save()

            wb = load_workbook("./poll_results/" + keyQuizName + "_stats.xlsx")

            my_blue = openpyxl.styles.colors.Color(rgb='CCE5FF')
            my_fill_blue = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_blue)
            my_red = openpyxl.styles.colors.Color(rgb='FFCCCC')
            my_fill_red = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            my_green = openpyxl.styles.colors.Color(rgb='CCFFCC')
            my_fill_green = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
            my_orange = openpyxl.styles.colors.Color(rgb='FFE5CC')
            my_fill_orange = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_orange)


            for k,ws in enumerate(wb.worksheets):

                correctAnswer=""
                for j, quizPart in enumerate(quizStat.questionStatDict):
                    str2=str(ws['A1'].value)
                    qmatch= StringComparator(quizPart.getQuestion().getQuestionText(),str2).cmp_ig_CaseSpacePunc
                    if qmatch==0:
                        correctAnswer = quizPart.getQuestion().getAnswer()
                        break


                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                                                   wrap_text=True, shrink_to_fit=False, indent=1)

                for cell in ws['A']:
                    cell.alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True,
                                               shrink_to_fit=False, indent=1)

                for column in ws.iter_cols():
                    column[0].fill = my_fill_blue

                for row in ws.iter_rows():
                    row[0].fill = my_fill_orange

                trueCellIndex = 0
                for cell in ws['A']:

                    s = cell.internal_value
                    if s is not None:

                        for answer in correctAnswer:

                            match = StringComparator(str(s), answer).cmp_ig_C_S_P_N
                            if match == 0:
                                cell.fill = my_fill_green
                                trueCellIndex = cell.row
                                break

                        if match!=0:
                            cell.fill = my_fill_red

                ws['A1'].fill = my_fill_orange
                ws['A1'].font = Font(bold=True)
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['A'].height = 20
                p = BarChart3D()
                data = Reference(worksheet=ws, min_col=2, max_col=ws.max_column, min_row=2, max_row=ws.max_row)

                # p.add_data(data)
                p.title = "Answer Distribution"

                pt = DataPoint(idx=trueCellIndex - 2)
                pt.graphicalProperties.solidFill = "b2ff59"

                answers = Reference(worksheet=ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)

                l = list(ws.iter_rows())

                g = 2
                for row in ws.iter_rows():
                    value = Reference(worksheet=ws, min_col=3, max_col=3, min_row=g, max_row=g)

                    ans = str(ws['A' + str(g)].internal_value)

                    serie = Series(value, title=ans)

                    if ans != None:

                        for answer in correctAnswer:

                            match = StringComparator(ans, answer).cmp_ig_C_S_P_N
                            if match == 0:
                                serie.graphicalProperties.solidFill = "b2ff59"
                            break

                    p.append(serie)
                    g += 1

                p.height = 13
                p.width = 20
                p.style = 2
                p.legend.legendPos.upper()
                p.y_axis.scaling.min = 0
                p.y_axis.scaling.max = 100
                p.y_axis.title = "%"
                ws.add_chart(p, "D1")
                k += 1
            wb.save("./poll_results/" + keyQuizName + "_stats.xlsx")








